import io
from datetime import date
from typing import Any

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

_BLANK = "." * 40

_KHMER_FONT = "Kantumruy Pro"
_KHMER_TITLE = Font(name=_KHMER_FONT, size=14, bold=True)
_KHMER_BOLD = Font(name=_KHMER_FONT, size=10, bold=True)
_KHMER_REGULAR = Font(name=_KHMER_FONT, size=10, bold=False)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center")
_THIN_SIDE = Side(border_style="thin", color="000000")
_FULL_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)

_DATA_COLUMNS = ("A", "B", "C", "D", "E", "F", "G", "H", "I")
_NUMERIC_COLUMNS = ("G",)

# Relative column widths as % of the table — proportions stay fixed while the
# absolute scale (_TABLE_WIDTH_UNITS) can be tuned to fit the printable page.
_COLUMN_WIDTH_PERCENT = {"A": 6, "B": 13, "C": 13, "D": 15, "E": 16, "F": 13, "G": 10, "H": 10, "I": 14}
_TABLE_WIDTH_UNITS = 120  # total width, in Excel column-width units, at 100%


def _column_widths() -> dict[str, float]:
    return {col: round(pct / 100 * _TABLE_WIDTH_UNITS, 1) for col, pct in _COLUMN_WIDTH_PERCENT.items()}


def _header_value(value: Any) -> str:
    return str(value) if value not in (None, "") else _BLANK


def _format_period(date_from: date | None, date_to: date | None) -> str:
    if date_from and date_to:
        return f"{date_from.strftime('%d/%m/%Y')} - {date_to.strftime('%d/%m/%Y')}"
    return _BLANK


def _write_titles(
    ws: Worksheet,
    representative: str | None,
    date_from: date | None,
    date_to: date | None,
    report_date: date,
) -> None:
    ws["B1"] = "របាយការណ៍ប្រវត្តិត្រលប់ថ្នាំជក់"
    ws["B1"].font = _KHMER_TITLE
    ws["B1"].alignment = _CENTER
    ws.merge_cells("B1:H3")

    fields = {
        "A4": f"អ្នកតំណាង : {_header_value(representative)}",
        "F4": f"ថ្ងៃបង្កើត : {report_date.strftime('%d/%m/%Y')}",
        "A5": f"កាលបរិច្ឆេទ : {_format_period(date_from, date_to)}",
    }
    for cell_ref, text in fields.items():
        ws[cell_ref] = text
        ws[cell_ref].font = _KHMER_BOLD
        ws[cell_ref].alignment = _LEFT


def _write_table_headers(ws: Worksheet) -> None:
    headers = {
        "A8": "ល.រ",
        "B8": "លេខបង្គរ\nRepay No",
        "C8": "កុងត្រាលេខ\nContract No",
        "D8": "អ្នកតំណាង\nRepresentative",
        "E8": "កសិករ\nFarmer",
        "F8": "ប្រភេទថ្នាំ\nTobacco",
        "G8": "ចំនួនគីឡូ\nQty Kg",
        "H8": "ថ្ងៃខែ\nDate",
        "I8": "សម្គាល់\nRemark",
    }
    for cell_ref, text in headers.items():
        ws[cell_ref] = text
        ws[cell_ref].font = _KHMER_BOLD
        ws[cell_ref].alignment = _CENTER

        col_letter = cell_ref[0]
        ws.merge_cells(f"{col_letter}8:{col_letter}9")
        ws[f"{col_letter}8"].border = _FULL_BORDER
        ws[f"{col_letter}9"].border = _FULL_BORDER


def _write_data_row(ws: Worksheet, row_idx: int, row_no: int, row: dict[str, Any]) -> None:
    values: dict[str, Any] = {
        "A": row_no,
        "B": row.get("repay_num"),
        "C": row.get("con_num"),
        "D": row.get("representative"),
        "E": row.get("farmer_name"),
        "F": row.get("tobacco_type"),
        "G": row.get("qty_repay"),
        "H": row.get("repay_date"),
        "I": row.get("note"),
    }
    ws.row_dimensions[row_idx].height = 30
    for col_letter in _DATA_COLUMNS:
        cell = ws[f"{col_letter}{row_idx}"]
        cell.value = values[col_letter]
        cell.font = _KHMER_REGULAR
        cell.alignment = _CENTER
        cell.border = _FULL_BORDER
        if col_letter in _NUMERIC_COLUMNS:
            cell.number_format = "#,##0.00"
        elif col_letter == "H":
            cell.number_format = "dd/mm/yyyy"


def _write_totals_row(ws: Worksheet, row_idx: int, rows: list[dict[str, Any]]) -> None:
    ws[f"F{row_idx}"] = "សរុប / Total"
    ws[f"G{row_idx}"] = sum(r.get("qty_repay") or 0 for r in rows)
    for col_letter in ("F", "G"):
        cell = ws[f"{col_letter}{row_idx}"]
        cell.font = _KHMER_BOLD
        cell.border = _FULL_BORDER
        if col_letter == "G":
            cell.number_format = "#,##0.00"


def build_tobacco_repay_template(
    *,
    representative: str | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    report_date: date | None = None,
    rows: list[dict[str, Any]] | None = None,
) -> io.BytesIO:
    """Build the tobacco-repay history report as an .xlsx stream.

    `rows` holds one dict per repay record (repay_num, con_num, representative,
    farmer_name, tobacco_type, qty_repay, repay_date, note); header fields fall
    back to a blank, fillable line when not supplied.
    """
    report_date = report_date or date.today()
    rows = rows or []

    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    ws.title = "Repay History"
    ws.views.sheetView[0].showGridLines = True

    ws.page_margins = PageMargins(left=1, right=1, top=1, bottom=1)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    _write_titles(ws, representative, date_from, date_to, report_date)
    _write_table_headers(ws)

    first_data_row = 10
    for offset, row in enumerate(rows):
        _write_data_row(ws, first_data_row + offset, offset + 1, row)

    if rows:
        _write_totals_row(ws, first_data_row + len(rows), rows)

    for col_letter, width in _column_widths().items():
        ws.column_dimensions[col_letter].width = width

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
